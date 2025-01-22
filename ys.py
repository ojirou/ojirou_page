import requests
import bs4
import pandas as pd
import subprocess
import os
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.styles.colors import Color
import lxml.html
# YAHOO!JAPAN検索をする関数
def search_keyword_yahoo(keyword, start_num):
    searched_list = []
    rank_num = start_num + 1
    url = f'http://search.yahoo.co.jp/search?p={keyword}&ei=UTF-8'
    if start_num != 0:
        url += f'&b={start_num}'
    
    req = requests.get(url)
    html = req.text
    dom = lxml.html.fromstring(html)
    a_list = dom.xpath("//div[@id='web']//li/a")

    for i, j in enumerate(a_list):
        page_data = []
        page_data.append(rank_num)
        page_data.append(j.text_content())
        page_data.append(j.attrib['href'])
        searched_list.append(page_data)
        rank_num += 1
    req.close()
    return searched_list

def main():
    Dir = r'C:\\Users\\user\\python\\web_scraping\\YahooSearch\\'
    search_keyword = input('キーワード >> ')
    search_keyword2 = search_keyword.replace(" ", "_")
    ExcelName = Dir + search_keyword2 + '.xlsx'

    # 検索結果取得
    results = []
    b_list = [0, 11, 22, 33, 44, 55, 66, 77, 88, 99, 110, 121, 132, 143, 154, 165, 176, 187, 198, 209]
    for b in b_list:
        results.extend(search_keyword_yahoo(search_keyword, b))

    # データフレーム作成
    df = pd.DataFrame(results, columns=['順位', 'タイトル', 'URL'])
    df.to_excel(ExcelName, index=False)

    # エクセル操作
    FileName = ExcelName
    wb = load_workbook(FileName)
    ws = wb.active

    # URL列の幅とハイパーリンク設定
    ws.column_dimensions['B'].width = 60
    ws.column_dimensions['C'].width = 100
    column_num = 3
    row_nums = len(df)
    for row_num in range(1, row_nums + 2):
        target_cell = ws.cell(row=row_num, column=column_num)
        Address = target_cell.value
        target_cell.value = f'=HYPERLINK("{Address}", "{Address}")'
        target_cell.font = Font(size=9, color=Color(theme=10))

    wb.save(FileName)

    # ファイルを開く
    subprocess.Popen(['start', FileName], shell=True)
    os.startfile(FileName)
    os.startfile(Dir, operation='open')

if __name__ == "__main__":
    main()
