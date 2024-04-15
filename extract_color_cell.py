import os
import openpyxl
import subprocess
# 書き込み用のxlsxファイルを作成
wb_output = openpyxl.Workbook()
ws_temporary = wb_output.create_sheet("temporary")
ws_result = wb_output.create_sheet("抽出結果")
# ファイル名の指定
output_filename = r"C:\Users\user\python\EXCEL\240414マーキング記事抽出.xlsx"
# 指定フォルダ内のすべてのxlsxファイルに対して処理を実行
folder_path = r"C:\Users\user\python\web_scraping\GoogleSearch\test"
for file_name in os.listdir(folder_path):
    if file_name.endswith(".xlsx"):
        # ファイルを開く
        print(folder_path)
        print(file_name)
        wb_input = openpyxl.load_workbook(os.path.join(folder_path, file_name))
        ws_input = wb_input.active
        # 抽出結果シートに元ファイル名を記入
        # 2行目から行末まで処理を実行
        for row in ws_input.iter_rows(min_row=2):
            if row[1].fill.start_color.index != '00000000':  # 透明な場合
                ws_result.append([file_name])
                ws_temporary.append([cell.value for cell in row])
        # ファイルを保存
        wb_input.close()
# 書き込み用xlsxファイルを保存
wb_output.save(output_filename)
subprocess.Popen(["start", "", output_filename], shell=True)